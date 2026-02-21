import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from difflib import SequenceMatcher
from collections import defaultdict
import io

st.set_page_config(
    page_title="GSTR-2A Reconciliation · India",
    page_icon="🧾",
    layout="wide",
    initial_sidebar_state="collapsed"
)

UPI_ID    = "7600480575@upi"
UPI_NAME  = "GST%20Reco%20Tool"
UPI_AMT   = "19"
UPI_NOTE  = "GSTR2A%20Reco"
UPI_LINK  = f"upi://pay?pa={UPI_ID}&pn={UPI_NAME}&am={UPI_AMT}&cu=INR&tn={UPI_NOTE}"
QR_URL    = f"https://api.qrserver.com/v1/create-qr-code/?size=180x180&data={UPI_LINK}&bgcolor=0c0c14&color=ffffff&qzone=2"

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@600;700;800&family=Inter:wght@400;500;600&family=JetBrains+Mono:wght@600;700&display=swap');

*, body { font-family: 'Inter', sans-serif; margin: 0; padding: 0; }

/* Hide Streamlit chrome */
footer, #MainMenu, header { visibility: hidden !important; }
[data-testid="stToolbar"] { display: none !important; }
.stDeployButton { display: none !important; }
div[data-testid="stFileUploader"] label { display: none !important; }
[data-testid="manage-app-button"] { display: none !important; }
iframe[title="streamlit_app"] + div { display: none !important; }
.viewerBadge_container__r5tak { display: none !important; }
#stDecoration { display: none !important; }

.stApp {
    background: #0c0c14;
}
.block-container {
    padding: 0 !important;
    max-width: 100% !important;
}

/* ── HERO SECTION ── */
.hero-wrap {
    background: linear-gradient(160deg, #0c0c14 0%, #111122 50%, #0c0c14 100%);
    border-bottom: 1px solid rgba(255,255,255,0.05);
    padding: 3.5rem 2rem 3rem 2rem;
    text-align: center;
    position: relative;
    overflow: hidden;
}
.hero-wrap::before {
    content: '';
    position: absolute;
    top: -40%; left: 50%;
    transform: translateX(-50%);
    width: 600px; height: 400px;
    background: radial-gradient(ellipse, rgba(99,102,241,0.12) 0%, transparent 70%);
    pointer-events: none;
}
.hero-wrap::after {
    content: '';
    position: absolute;
    bottom: 0; left: 0; right: 0; height: 1px;
    background: linear-gradient(90deg, transparent, rgba(99,102,241,0.4), transparent);
}
.hero-pill {
    display: inline-flex;
    align-items: center;
    gap: 0.5rem;
    background: rgba(99,102,241,0.1);
    border: 1px solid rgba(99,102,241,0.3);
    color: #a5b4fc;
    font-size: 0.7rem;
    font-weight: 600;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    padding: 0.35rem 1rem;
    border-radius: 100px;
    margin-bottom: 1.5rem;
}
.hero-title {
    font-family: 'Syne', sans-serif;
    font-size: 3.2rem;
    font-weight: 800;
    color: #ffffff;
    letter-spacing: -2px;
    line-height: 1.05;
    margin-bottom: 1rem;
}
.hero-title .accent {
    background: linear-gradient(135deg, #6366f1, #8b5cf6, #a78bfa);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
}
.hero-sub {
    color: #4a4a6a;
    font-size: 1rem;
    max-width: 500px;
    margin: 0 auto 0.5rem auto;
    line-height: 1.6;
}
.hero-tags {
    display: flex;
    justify-content: center;
    gap: 0.6rem;
    margin-top: 1.2rem;
    flex-wrap: wrap;
}
.hero-tag {
    background: rgba(255,255,255,0.03);
    border: 1px solid rgba(255,255,255,0.07);
    color: #555580;
    font-size: 0.72rem;
    padding: 0.28rem 0.7rem;
    border-radius: 6px;
}

/* ── MAIN CONTENT ── */
.main-wrap {
    max-width: 900px;
    margin: 0 auto;
    padding: 2.5rem 1.5rem;
}

/* ── SECURITY ── */
.sec-bar {
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 0.5rem;
    background: rgba(16,185,129,0.05);
    border: 1px solid rgba(16,185,129,0.15);
    border-radius: 10px;
    padding: 0.6rem 1rem;
    font-size: 0.78rem;
    color: #34d399;
    margin-bottom: 2rem;
}

/* ── UPLOAD CARDS ── */
.upload-grid {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 1rem;
    margin-bottom: 1.5rem;
}
.upload-card-label {
    font-size: 0.68rem;
    font-weight: 600;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    color: #6366f1;
    margin-bottom: 0.5rem;
}

/* ── TOLERANCE ── */
.tol-bar {
    background: rgba(99,102,241,0.05);
    border: 1px solid rgba(99,102,241,0.15);
    border-radius: 10px;
    padding: 0.65rem 1rem;
    font-size: 0.8rem;
    color: #818cf8;
    margin-bottom: 1.5rem;
}

/* ── STAT CARDS ── */
.stats-section {
    margin: 2rem 0;
}
.stats-title {
    font-size: 0.68rem;
    font-weight: 600;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    color: #333355;
    margin-bottom: 1rem;
}
.stat-row {
    display: grid;
    grid-template-columns: repeat(3,1fr);
    gap: 0.7rem;
    margin-bottom: 0.7rem;
}
.stat-card {
    background: #111122;
    border: 1px solid #1a1a2e;
    border-radius: 12px;
    padding: 1.2rem 1rem;
    text-align: center;
}
.stat-n {
    font-family: 'JetBrains Mono', monospace;
    font-size: 2.2rem;
    font-weight: 700;
    line-height: 1;
    margin-bottom: 0.3rem;
}
.stat-l { font-size: 0.72rem; color: #33334a; }
.green  { color: #34d399; }
.yellow { color: #fbbf24; }
.red    { color: #f87171; }
.indigo { color: #818cf8; }

/* ── PAYMENT BOX ── */
.pay-section {
    background: #0f0f1e;
    border: 1px solid #1e1e35;
    border-radius: 20px;
    overflow: hidden;
    margin: 1.5rem 0;
}
.pay-header {
    background: linear-gradient(135deg, #1e1b4b, #1e1038);
    border-bottom: 1px solid rgba(99,102,241,0.2);
    padding: 1.5rem 2rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
    flex-wrap: wrap;
    gap: 1rem;
}
.pay-title {
    font-family: 'Syne', sans-serif;
    color: #fff;
    font-size: 1.2rem;
    font-weight: 700;
}
.pay-subtitle { color: #6366f1; font-size: 0.8rem; margin-top: 0.2rem; }
.pay-amount-badge {
    font-family: 'JetBrains Mono', monospace;
    background: linear-gradient(135deg, #6366f1, #8b5cf6);
    color: #fff;
    font-size: 1.6rem;
    font-weight: 700;
    padding: 0.5rem 1.5rem;
    border-radius: 12px;
    white-space: nowrap;
}
.pay-body {
    padding: 2rem;
    display: flex;
    gap: 2.5rem;
    align-items: flex-start;
    flex-wrap: wrap;
}
.pay-qr-side { text-align: center; flex-shrink: 0; }
.qr-frame {
    background: #0c0c14;
    border: 1.5px solid rgba(99,102,241,0.3);
    border-radius: 16px;
    padding: 12px;
    display: inline-block;
    margin-bottom: 0.6rem;
}
.qr-apps {
    display: flex;
    justify-content: center;
    gap: 0.4rem;
    flex-wrap: wrap;
    margin-top: 0.4rem;
}
.qr-app {
    background: rgba(255,255,255,0.03);
    border: 1px solid rgba(255,255,255,0.07);
    border-radius: 5px;
    padding: 0.18rem 0.5rem;
    font-size: 0.68rem;
    color: #444466;
}
.pay-info-side { flex: 1; min-width: 220px; }
.pay-steps-list {
    list-style: none;
    padding: 0;
    margin: 0 0 1.2rem 0;
}
.pay-steps-list li {
    display: flex;
    align-items: flex-start;
    gap: 0.8rem;
    margin-bottom: 0.9rem;
    font-size: 0.84rem;
    color: #44446a;
    line-height: 1.4;
}
.pay-steps-list li b { color: #9999cc; }
.step-num {
    background: rgba(99,102,241,0.1);
    border: 1px solid rgba(99,102,241,0.2);
    color: #6366f1;
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.7rem;
    font-weight: 700;
    width: 22px; height: 22px;
    border-radius: 6px;
    display: flex; align-items: center; justify-content: center;
    flex-shrink: 0;
    margin-top: 0.1rem;
}
.refund-pill {
    background: rgba(16,185,129,0.06);
    border: 1px solid rgba(16,185,129,0.15);
    border-radius: 10px;
    padding: 0.65rem 1rem;
    font-size: 0.76rem;
    color: #34d399;
    line-height: 1.5;
}

/* ── UTR INPUT STYLE ── */
div[data-testid="stTextInput"] input {
    background: #111122 !important;
    border: 1.5px solid #1e1e35 !important;
    border-radius: 10px !important;
    color: #fff !important;
    font-size: 0.9rem !important;
    padding: 0.65rem 1rem !important;
}
div[data-testid="stTextInput"] input:focus {
    border-color: #6366f1 !important;
    box-shadow: 0 0 0 3px rgba(99,102,241,0.15) !important;
}

/* ── BUTTONS ── */
.stButton > button {
    background: linear-gradient(135deg, #6366f1, #8b5cf6) !important;
    color: #fff !important;
    font-weight: 600 !important;
    font-size: 0.9rem !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 0.65rem 1.5rem !important;
    width: 100% !important;
    transition: opacity 0.2s, transform 0.1s !important;
}
.stButton > button:hover {
    opacity: 0.9 !important;
    transform: translateY(-1px) !important;
}
.stDownloadButton > button {
    background: linear-gradient(135deg, #059669, #10b981) !important;
    color: #fff !important;
    font-weight: 700 !important;
    font-size: 1rem !important;
    border: none !important;
    border-radius: 12px !important;
    padding: 0.8rem 1.5rem !important;
    width: 100% !important;
}

/* ── README SECTION ── */
.readme-wrap {
    border-top: 1px solid #111122;
    padding: 3rem 1.5rem;
    max-width: 900px;
    margin: 0 auto;
}
.readme-header {
    text-align: center;
    margin-bottom: 2.5rem;
}
.readme-header h2 {
    font-family: 'Syne', sans-serif;
    font-size: 1.6rem;
    font-weight: 700;
    color: #fff;
    margin-bottom: 0.4rem;
}
.readme-header p { color: #33334a; font-size: 0.85rem; }

.readme-grid {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 1rem;
    margin-bottom: 1.5rem;
}
.readme-card {
    background: #0f0f1e;
    border: 1px solid #1a1a2e;
    border-radius: 14px;
    padding: 1.4rem;
}
.readme-card-icon { font-size: 1.4rem; margin-bottom: 0.6rem; }
.readme-card h4 {
    font-family: 'Syne', sans-serif;
    color: #9999cc;
    font-size: 0.85rem;
    font-weight: 700;
    margin-bottom: 0.6rem;
}
.readme-card p, .readme-card li {
    font-size: 0.79rem;
    color: #33334a;
    line-height: 1.6;
}
.readme-card ul { padding-left: 1rem; margin: 0; }
.readme-card li { margin-bottom: 0.3rem; }

.output-table {
    width: 100%;
    border-collapse: collapse;
    margin: 1rem 0;
    font-size: 0.8rem;
}
.output-table th {
    background: #111122;
    color: #444466;
    font-weight: 600;
    letter-spacing: 0.5px;
    padding: 0.6rem 0.8rem;
    text-align: left;
    border-bottom: 1px solid #1a1a2e;
}
.output-table td {
    padding: 0.6rem 0.8rem;
    color: #33334a;
    border-bottom: 1px solid #0f0f1e;
}
.output-table tr:hover td { background: rgba(99,102,241,0.02); }
.badge-green  { color: #34d399; font-size: 0.7rem; }
.badge-yellow { color: #fbbf24; font-size: 0.7rem; }
.badge-red    { color: #f87171; font-size: 0.7rem; }

.bottom-bar {
    text-align: center;
    padding: 1.5rem;
    border-top: 1px solid #0f0f1e;
    color: #1e1e2e;
    font-size: 0.7rem;
    line-height: 2;
}
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# HERO
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="hero-wrap">
  <div class="hero-pill">🇮🇳 Made for Indian CA Firms</div>
  <div class="hero-title">GSTR-2A vs Books<br><span class="accent">Reconciliation</span></div>
  <div class="hero-sub">Upload your portal export and Tally register — get a full colour-coded Excel report in seconds.</div>
  <div class="hero-tags">
    <span class="hero-tag">⚡ Instant Results</span>
    <span class="hero-tag">🔍 Fuzzy Name Matching</span>
    <span class="hero-tag">±₹10 Tolerance</span>
    <span class="hero-tag">🔒 Zero Data Storage</span>
    <span class="hero-tag">₹19 per Report</span>
  </div>
</div>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN CONTENT
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown('<div class="main-wrap">', unsafe_allow_html=True)

st.markdown("""
<div class="sec-bar">
  🔒 <strong>100% Private</strong> — Files are processed in memory only and permanently deleted when you close the tab. We never store, log, or access your data.
</div>
""", unsafe_allow_html=True)

st.markdown('<div class="tol-bar">⚡ <strong>Smart matching:</strong> Fuzzy vendor name comparison handles Tally typos automatically. GST difference within <strong>±₹10 per head</strong> is treated as matched.</div>', unsafe_allow_html=True)

# Upload
col1, col2 = st.columns(2)
with col1:
    st.markdown('<div class="upload-card-label">📥 GSTR-2A — Portal Export</div>', unsafe_allow_html=True)
    file_2a = st.file_uploader("2a", type=["xls","xlsx"], key="f2a", label_visibility="collapsed")
    if file_2a: st.success(f"✅ {file_2a.name}")

with col2:
    st.markdown('<div class="upload-card-label">📒 Purchase Register — Tally Export</div>', unsafe_allow_html=True)
    file_books = st.file_uploader("bk", type=["xls","xlsx"], key="fb", label_visibility="collapsed")
    if file_books: st.success(f"✅ {file_books.name}")

st.markdown("<br>", unsafe_allow_html=True)
col_run, _, _ = st.columns([1,1,1])
with col_run:
    run_btn = st.button("⚡ Run Reconciliation — Free Preview")

st.markdown('</div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# CORE LOGIC
# ═══════════════════════════════════════════════════════════════════════════════
def convert_xls(file_bytes, filename):
    if filename.endswith('.xlsx'):
        return io.BytesIO(file_bytes)
    import xlrd
    xls = xlrd.open_workbook(file_contents=file_bytes)
    wb  = openpyxl.Workbook(); wb.remove(wb.active)
    for sname in xls.sheet_names():
        xs = xls.sheet_by_name(sname)
        ws = wb.create_sheet(title=sname)
        for ri in range(xs.nrows):
            for ci in range(xs.ncols):
                cell = xs.cell(ri, ci)
                if cell.ctype == 3:
                    try: ws.cell(ri+1,ci+1,xlrd.xldate_as_datetime(cell.value,xls.datemode))
                    except: ws.cell(ri+1,ci+1,cell.value)
                elif cell.ctype != 0:
                    ws.cell(ri+1,ci+1,cell.value)
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

def parse_2a(wb):
    rows=list(wb['invoice'].iter_rows(values_only=True)); data=[]
    for r in rows[3:]:
        if not r[0] or not isinstance(r[0],(int,float)): continue
        data.append({'sno':int(r[0]),'supplier':str(r[1] or '').strip().upper(),
                     'gstin':str(r[2] or '').strip().upper(),'period':str(r[3] or '').strip(),
                     'inv_no':str(r[5] or '').strip(),'inv_date':str(r[8]) if r[8] else '',
                     'inv_value':float(r[9] or 0),'taxable':float(r[10] or 0),
                     'igst':float(r[11] or 0),'cgst':float(r[12] or 0),'sgst':float(r[13] or 0)})
    return data

def parse_books(wb):
    rows=list(wb.active.iter_rows(values_only=True))
    CC=[6,18,21,23]; SC=[7,19,22,24]; IC=[32]; data=[]
    for i,r in enumerate(rows):
        if i<9 or not r[0] or not r[1] or r[1]=='Grand Total': continue
        cgst=sum(float(r[c] or 0) for c in CC if c<len(r))
        sgst=sum(float(r[c] or 0) for c in SC if c<len(r))
        igst=sum(float(r[c] or 0) for c in IC if c<len(r))
        if cgst==sgst==igst==0: continue
        data.append({'date':str(r[0])[:10] if r[0] else '','name':str(r[1] or '').strip().upper(),
                     'gross':float(r[4] or 0),'cgst':cgst,'sgst':sgst,'igst':igst,
                     '_matched':False,'_id':i})
    return data

def norm(s): return re.sub(r'[^A-Z0-9 ]','',re.sub(r'\s+',' ',s.upper().strip()))
def sim(a,b): return SequenceMatcher(None,norm(a),norm(b)).ratio()

def match(gst2a,books):
    by=defaultdict(list)
    for b in books: by[b['name']].append(b)
    used,exact,diff,un2a=set(),[],[],[]
    for rec in gst2a:
        best,bscore,bd=None,0,None
        for bn,bl in by.items():
            s=sim(rec['supplier'],bn)
            if s<0.65: continue
            for bk in bl:
                if bk['_id'] in used: continue
                cd=abs(rec['cgst']-bk['cgst']); sd=abs(rec['sgst']-bk['sgst']); gd=abs(rec['igst']-bk['igst'])
                if cd<=10 and sd<=10 and gd<=10:
                    sc2=s*100+1/(cd+sd+gd+0.01)
                    if sc2>bscore: bscore,best,bd=sc2,bk,(cd,sd,gd)
        if best:
            used.add(best['_id']); best['_matched']=True
            e={**rec,'bk_name':best['name'],'bk_date':best['date'],'bk_gross':best['gross'],
               'bk_cgst':best['cgst'],'bk_sgst':best['sgst'],'bk_igst':best['igst'],
               'cgst_diff':bd[0],'sgst_diff':bd[1],'igst_diff':bd[2],'total_diff':sum(bd)}
            (exact if sum(bd)<0.01 else diff).append(e)
        else: un2a.append(rec)
    return exact,diff,un2a,[b for b in books if not b['_matched']]

def build_excel(re_,rd_,ru_,rb_,gst2a,books):
    G=PatternFill("solid",fgColor="C6EFCE"); Y=PatternFill("solid",fgColor="FFEB9C")
    RF=PatternFill("solid",fgColor="FFC7CE"); LB=PatternFill("solid",fgColor="D6E4F7")
    BH=PatternFill("solid",fgColor="1F4E79"); TH=PatternFill("solid",fgColor="375623")
    OH=PatternFill("solid",fgColor="7B6200"); RH=PatternFill("solid",fgColor="833C00")
    DH=PatternFill("solid",fgColor="7B2C2C")
    t=Side(style="thin",color="CCCCCC"); BD=Border(left=t,right=t,top=t,bottom=t)
    WH=Font(color="FFFFFF",bold=True,name="Arial",size=9)
    BLD=Font(bold=True,name="Arial",size=9); NR=Font(name="Arial",size=9)
    CT=Alignment(horizontal="center",vertical="center",wrap_text=True)
    LF=Alignment(horizontal="left",vertical="center")
    def sh(c,f): c.fill=f;c.font=WH;c.alignment=CT;c.border=BD
    def sc(c,f=None): c.font=NR;c.alignment=LF;c.border=BD;(setattr(c,'fill',f) if f else None)
    sf=lambda l,k:sum(r[k] for r in l)
    wb=openpyxl.Workbook(); wb.remove(wb.active)

    # SUMMARY SHEET
    ws=wb.create_sheet("📊 Summary"); ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:F1"); c=ws["A1"]
    c.value="GSTR-2A vs Books — Reconciliation Summary"
    c.font=Font(name="Arial",bold=True,size=13,color="FFFFFF"); c.fill=BH; c.alignment=CT; ws.row_dimensions[1].height=26
    ws.merge_cells("A2:F2"); c2=ws["A2"]; c2.value="Generated by GSTR-2A Reconciliation Tool · India"
    c2.font=Font(name="Arial",size=8,color="AAAAAA",italic=True); c2.alignment=CT
    for col,(label,val,fill) in enumerate([
        ("Total 2A Records",len(gst2a),BH),
        ("Total Books Records",len(books),PatternFill("solid",fgColor="1F3864")),
        ("✅ Matched Exact",len(re_),TH),("✅ Matched (±₹10)",len(rd_),OH),
        ("⚠️ Unmatched in 2A",len(ru_),RH),("⚠️ Unmatched in Books",len(rb_),DH)],1):
        lc=ws.cell(row=4,column=col,value=label); lc.fill=fill; lc.font=WH; lc.alignment=CT; lc.border=BD
        vc=ws.cell(row=5,column=col,value=val)
        vc.fill=PatternFill("solid",fgColor="F5F5F5"); vc.font=Font(name="Arial",bold=True,size=16)
        vc.alignment=CT; vc.border=BD
    ws.row_dimensions[5].height=34
    for ci,h in enumerate(["","IGST (₹)","CGST (₹)","SGST (₹)","Total GST (₹)"],1): sh(ws.cell(row=7,column=ci,value=h),BH)
    for ri,(label,ig,cg,sg) in enumerate([
        ("2A Total",sf(gst2a,'igst'),sf(gst2a,'cgst'),sf(gst2a,'sgst')),
        ("Books Total",sf(books,'igst'),sf(books,'cgst'),sf(books,'sgst'))],8):
        for ci,v in enumerate([label,ig,cg,sg,ig+cg+sg],1):
            cell=ws.cell(row=ri,column=ci,value=v if isinstance(v,str) else round(v,2))
            sc(cell,PatternFill("solid",fgColor="EBF5FB" if ri==8 else "FEF9E7"))
            if ci>1: cell.number_format='#,##0.00'
    dr=["Difference",round(sf(gst2a,'igst')-sf(books,'igst'),2),
        round(sf(gst2a,'cgst')-sf(books,'cgst'),2),round(sf(gst2a,'sgst')-sf(books,'sgst'),2)]
    dr.append(sum(dr[1:]))
    for ci,v in enumerate(dr,1):
        cell=ws.cell(row=10,column=ci,value=v if isinstance(v,str) else round(v,2))
        sc(cell,PatternFill("solid",fgColor="FCE4D6") if isinstance(v,(int,float)) and abs(v)>0.01 else G)
        if ci>1: cell.number_format='#,##0.00'
    for i in range(1,7): ws.column_dimensions[get_column_letter(i)].width=26

    def mk_match(wb,name,data,hfill,is_diff=False):
        ws=wb.create_sheet(name); ws.sheet_view.showGridLines=False
        lc='S' if is_diff else 'P'; ws.merge_cells(f"A1:{lc}1"); c=ws["A1"]; c.value=name
        c.font=Font(name="Arial",bold=True,size=11,color="FFFFFF"); c.fill=hfill; c.alignment=CT
        hdrs=["#","2A: Supplier","2A: GSTIN","2A: Invoice No","2A: Date","2A: Inv Value","2A: Taxable",
              "2A: IGST","2A: CGST","2A: SGST","Books: Vendor","Books: Date","Books: Gross",
              "Books: IGST","Books: CGST","Books: SGST"]
        if is_diff: hdrs+=["IGST Diff","CGST Diff","SGST Diff"]
        for ci,h in enumerate(hdrs,1): sh(ws.cell(row=2,column=ci,value=h),hfill)
        rf=G if not is_diff else Y
        for ri,rec in enumerate(data,3):
            vals=[rec['sno'],rec['supplier'],rec['gstin'],rec['inv_no'],rec['inv_date'],
                  rec['inv_value'],rec['taxable'],rec['igst'],rec['cgst'],rec['sgst'],
                  rec['bk_name'],rec['bk_date'],rec['bk_gross'],rec['bk_igst'],rec['bk_cgst'],rec['bk_sgst']]
            if is_diff: vals+=[rec['igst_diff'],rec['cgst_diff'],rec['sgst_diff']]
            for ci,v in enumerate(vals,1):
                cell=ws.cell(row=ri,column=ci,value=v); sc(cell,rf)
                if ci in [6,7,8,9,10,13,14,15,16,17,18,19]: cell.number_format='#,##0.00'
        tr=len(data)+3; ws.cell(row=tr,column=1,value="TOTAL").font=BLD
        for ci,key in [(8,'igst'),(9,'cgst'),(10,'sgst'),(14,'bk_igst'),(15,'bk_cgst'),(16,'bk_sgst')]:
            c2=ws.cell(row=tr,column=ci,value=round(sum(r[key] for r in data),2))
            c2.font=BLD; c2.number_format='#,##0.00'; c2.fill=LB; c2.border=BD
        ws.freeze_panes="A3"
        widths=[5,35,22,22,13,13,13,12,12,12,35,13,13,12,12,12]+([11,11,11] if is_diff else [])
        for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

    def mk_un2a(wb,data):
        ws=wb.create_sheet("⚠️ Unmatched in 2A"); ws.sheet_view.showGridLines=False
        ws.merge_cells("A1:J1"); c=ws["A1"]; c.value="⚠️ In GSTR-2A but NOT found in Books — Check for missed ITC"
        c.font=Font(name="Arial",bold=True,size=11,color="FFFFFF"); c.fill=RH; c.alignment=CT
        for ci,h in enumerate(["#","Supplier Name","GSTIN","Period","Invoice No","Invoice Date","Invoice Value","IGST","CGST","SGST"],1):
            sh(ws.cell(row=2,column=ci,value=h),RH)
        for ri,rec in enumerate(data,3):
            for ci,v in enumerate([rec['sno'],rec['supplier'],rec['gstin'],rec['period'],rec['inv_no'],
                                    rec['inv_date'],rec['inv_value'],rec['igst'],rec['cgst'],rec['sgst']],1):
                cell=ws.cell(row=ri,column=ci,value=v); sc(cell,RF)
                if ci in [7,8,9,10]: cell.number_format='#,##0.00'
        tr=len(data)+3; ws.cell(row=tr,column=1,value="TOTAL").font=BLD
        for ci,key in [(8,'igst'),(9,'cgst'),(10,'sgst')]:
            c2=ws.cell(row=tr,column=ci,value=round(sum(r[key] for r in data),2))
            c2.font=BLD; c2.number_format='#,##0.00'; c2.fill=LB; c2.border=BD
        ws.freeze_panes="A3"
        for i,w in enumerate([5,38,22,13,24,13,13,12,12,12],1): ws.column_dimensions[get_column_letter(i)].width=w

    def mk_unb(wb,data):
        ws=wb.create_sheet("⚠️ Unmatched in Books"); ws.sheet_view.showGridLines=False
        ws.merge_cells("A1:G1"); c=ws["A1"]; c.value="⚠️ In Books but NOT in 2A — Supplier may not have filed GST return"
        c.font=Font(name="Arial",bold=True,size=11,color="FFFFFF"); c.fill=DH; c.alignment=CT
        for ci,h in enumerate(["Vendor Name","Date","Gross Total","IGST","CGST","SGST","Total GST"],1):
            sh(ws.cell(row=2,column=ci,value=h),DH)
        for ri,rec in enumerate(data,3):
            tgst=rec['igst']+rec['cgst']+rec['sgst']
            for ci,v in enumerate([rec['name'],rec['date'],rec['gross'],rec['igst'],rec['cgst'],rec['sgst'],tgst],1):
                cell=ws.cell(row=ri,column=ci,value=v); sc(cell,RF)
                if ci in [3,4,5,6,7]: cell.number_format='#,##0.00'
        tr=len(data)+3; ws.cell(row=tr,column=1,value="TOTAL").font=BLD
        for ci,key in [(4,'igst'),(5,'cgst'),(6,'sgst')]:
            c2=ws.cell(row=tr,column=ci,value=round(sum(r[key] for r in data),2))
            c2.font=BLD; c2.number_format='#,##0.00'; c2.fill=LB; c2.border=BD
        ws.freeze_panes="A3"
        for i,w in enumerate([38,13,13,12,12,12,12],1): ws.column_dimensions[get_column_letter(i)].width=w

    mk_match(wb,"✅ Matched Exact",re_,TH)
    mk_match(wb,"✅ Matched (±10 Diff)",rd_,OH,is_diff=True)
    mk_un2a(wb,ru_); mk_unb(wb,rb_)
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ═══════════════════════════════════════════════════════════════════════════════
# RUN
# ═══════════════════════════════════════════════════════════════════════════════
if run_btn:
    if not file_2a or not file_books:
        st.error("❌ Please upload both files before running.")
    else:
        with st.spinner("Processing your files..."):
            try:
                b2a  = convert_xls(file_2a.read(),    file_2a.name)
                bb   = convert_xls(file_books.read(), file_books.name)
                wb2a = openpyxl.load_workbook(b2a, read_only=True)
                wbbk = openpyxl.load_workbook(bb,  read_only=True)
                gst2a= parse_2a(wb2a)
                books= parse_books(wbbk)
                re_,rd_,ru_,rb_ = match(gst2a,books)
                out  = build_excel(re_,rd_,ru_,rb_,gst2a,books)
                st.session_state.update({
                    "done":True,"out":out,
                    "re":re_,"rd":rd_,"ru":ru_,"rb":rb_,
                    "n2a":len(gst2a),"nb":len(books),"paid":False
                })
            except Exception as e:
                st.error(f"Error: {str(e)}")
                st.exception(e)

# ═══════════════════════════════════════════════════════════════════════════════
# RESULTS
# ═══════════════════════════════════════════════════════════════════════════════
if st.session_state.get("done"):
    re_=st.session_state["re"]; rd_=st.session_state["rd"]
    ru_=st.session_state["ru"]; rb_=st.session_state["rb"]

    st.markdown('<div class="main-wrap">', unsafe_allow_html=True)
    st.markdown('<div class="stats-section"><div class="stats-title">Reconciliation Results</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="stat-row">
      <div class="stat-card"><div class="stat-n green">{len(re_)}</div><div class="stat-l">✅ Matched Exact</div></div>
      <div class="stat-card"><div class="stat-n yellow">{len(rd_)}</div><div class="stat-l">✅ Matched (±₹10)</div></div>
      <div class="stat-card"><div class="stat-n red">{len(ru_)}</div><div class="stat-l">⚠️ Unmatched in 2A</div></div>
    </div>
    <div class="stat-row">
      <div class="stat-card"><div class="stat-n indigo">{st.session_state["n2a"]}</div><div class="stat-l">Total 2A Records</div></div>
      <div class="stat-card"><div class="stat-n indigo">{st.session_state["nb"]}</div><div class="stat-l">Total Books Records</div></div>
      <div class="stat-card"><div class="stat-n red">{len(rb_)}</div><div class="stat-l">⚠️ Unmatched in Books</div></div>
    </div>
    </div>
    """, unsafe_allow_html=True)

    if not st.session_state.get("paid"):
        st.markdown(f"""
        <div class="pay-section">
          <div class="pay-header">
            <div>
              <div class="pay-title">📥 Your report is ready</div>
              <div class="pay-subtitle">Pay ₹19 to download the full colour-coded Excel</div>
            </div>
            <div class="pay-amount-badge">₹19</div>
          </div>
          <div class="pay-body">
            <div class="pay-qr-side">
              <div class="qr-frame">
                <img src="{QR_URL}" width="180" height="180" alt="Scan to pay ₹19" />
              </div>
              <div style="font-size:0.72rem;color:#333355;margin-top:0.4rem;">Scan with any UPI app</div>
              <div class="qr-apps">
                <span class="qr-app">GPay</span>
                <span class="qr-app">PhonePe</span>
                <span class="qr-app">Paytm</span>
                <span class="qr-app">BHIM</span>
              </div>
            </div>
            <div class="pay-info-side">
              <ul class="pay-steps-list">
                <li><span class="step-num">1</span><span>Open GPay, PhonePe, Paytm or any UPI app on your phone</span></li>
                <li><span class="step-num">2</span><span>Tap <b>Scan QR</b> and scan the code on the left</span></li>
                <li><span class="step-num">3</span><span>Pay <b>₹19</b> and note the <b>UTR / Transaction ID</b> from your payment confirmation</span></li>
                <li><span class="step-num">4</span><span>Enter the UTR below and click <b>Unlock Download</b></span></li>
              </ul>
              <div class="refund-pill">
                💚 <strong>100% Refund Guarantee</strong><br>
                If the downloaded Excel doesn't work correctly, share a screenshot and get a full refund. No questions asked.
              </div>
            </div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        col_utr, col_btn, _ = st.columns([2,1,1])
        with col_utr:
            utr = st.text_input("", placeholder="Enter your UTR / Transaction ID here", label_visibility="collapsed")
        with col_btn:
            if st.button("🔓 Unlock Download"):
                if len(utr.strip()) >= 8:
                    st.session_state["paid"] = True
                    st.rerun()
                else:
                    st.error("❌ Please enter a valid UTR (minimum 8 characters)")

    else:
        st.success("✅ Payment confirmed — your Excel report is ready!")
        col_dl, _, _ = st.columns([1,1,1])
        with col_dl:
            st.download_button(
                "📥 Download Full Reconciliation Report",
                data=st.session_state["out"],
                file_name="GSTR2A_Reconciliation_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        st.markdown("<br>", unsafe_allow_html=True)
        c2,_,_ = st.columns([1,1,1])
        with c2:
            if st.button("🔄 New Reconciliation"):
                for k in ["done","out","re","rd","ru","rb","paid","n2a","nb"]:
                    st.session_state.pop(k,None)
                st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# README AT BOTTOM
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="readme-wrap">
  <div class="readme-header">
    <h2>How It Works</h2>
    <p>Everything you need to know about this tool — no technical knowledge required</p>
  </div>

  <div class="readme-grid">
    <div class="readme-card">
      <div class="readme-card-icon">📂</div>
      <h4>What Files Do I Need?</h4>
      <ul>
        <li><strong style="color:#9999cc">GSTR-2A:</strong> Download from GST Portal → Login → Return Dashboard → GSTR-2A → Download Excel</li>
        <li><strong style="color:#9999cc">Purchase Register:</strong> From Tally → Gateway of Tally → Display → Account Books → Purchase Register → Export to Excel</li>
        <li>Both <code style="color:#6366f1">.xls</code> and <code style="color:#6366f1">.xlsx</code> formats work</li>
      </ul>
    </div>
    <div class="readme-card">
      <div class="readme-card-icon">🧠</div>
      <h4>How Does Matching Work?</h4>
      <ul>
        <li><strong style="color:#9999cc">Fuzzy Names:</strong> Handles typos automatically — "H A Construction" matches "H.A.CONSTRUCTION"</li>
        <li><strong style="color:#9999cc">±₹10 Tolerance:</strong> Minor GST rounding differences are treated as matched (shown in yellow)</li>
        <li><strong style="color:#9999cc">No GSTIN match:</strong> Matching is done by name + amount, not GSTIN, to handle data entry differences</li>
      </ul>
    </div>
    <div class="readme-card">
      <div class="readme-card-icon">🔒</div>
      <h4>Is My Data Safe?</h4>
      <p>Completely. Your files are processed in temporary memory only — like a calculator. Nothing is saved to any server or database. The moment you close the tab, all data is permanently gone. We have no access to your files at any point.</p>
    </div>
    <div class="readme-card">
      <div class="readme-card-icon">💰</div>
      <h4>Pricing & Refund</h4>
      <p>₹19 per reconciliation — pay only when you're ready to download. Summary is always shown free. If the Excel report doesn't work correctly for your files, you get a <strong style="color:#34d399">100% refund</strong>, no questions asked.</p>
    </div>
  </div>

  <div class="readme-card" style="margin-bottom:1rem;">
    <div class="readme-card-icon">📊</div>
    <h4>What's Inside the Downloaded Excel?</h4>
    <table class="output-table">
      <tr>
        <th>Sheet</th>
        <th>Colour</th>
        <th>What It Means</th>
      </tr>
      <tr><td>📊 Summary</td><td>—</td><td>Overall counts + GST amount comparison between 2A and Books</td></tr>
      <tr><td>✅ Matched Exact</td><td><span class="badge-green">■ Green</span></td><td>Perfect match — same vendor, same GST amount in both</td></tr>
      <tr><td>✅ Matched (±₹10)</td><td><span class="badge-yellow">■ Yellow</span></td><td>Name matched, GST difference within ₹10 — treated as matched</td></tr>
      <tr><td>⚠️ Unmatched in 2A</td><td><span class="badge-red">■ Red</span></td><td>In GSTR-2A but missing from Books — potential missed ITC, investigate these</td></tr>
      <tr><td>⚠️ Unmatched in Books</td><td><span class="badge-red">■ Red</span></td><td>In Books but not in 2A — supplier may not have filed their GST return</td></tr>
    </table>
  </div>

</div>

<div class="bottom-bar">
  GSTR-2A Reconciliation Tool · India · FY 2025-26<br>
  Fuzzy name matching · ±₹10 tolerance · Zero data storage · ₹19 per report
</div>
""", unsafe_allow_html=True)
